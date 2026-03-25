"""Genera el archivo Excel con la tabla de diseño del benchmark."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()

# ── Colores y estilos ──────────────────────────────────────────────
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
exp_fills = {
    1: PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    2: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    3: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    4: PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
}
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ══════════════════════════════════════════════════════════════════════
# HOJA 1: Diseño de experimentos
# ══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Diseño de Experimentos"

headers = [
    "Experimento", "Enfoque", "Formato", "Enfoque de\nentrenamiento",
    "Punto de partida", "Entrada", "Batch", "Arquitectura",
    "Tarea", "Tamaños de\nmodelo", "RTX 5090", "Orin AGX", "Orin Nano",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

rows = [
    # Exp 1 — Comparación principal
    [1, "Comparación principal", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "Entren. + Inf.", "Inferencia", "Inferencia"],
    [1, "Comparación principal", "PyTorch FP32", "Transfer learning", "COCO preentrenado", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "Entren. + Inf.", "Inferencia", "Inferencia"],
    [1, "Comparación principal", "TensorRT FP16", "Desde cero", "Pesos aleatorios", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "—", "Export. + Inf.", "Export. + Inf."],
    [1, "Comparación principal", "TensorRT FP16", "Transfer learning", "COCO preentrenado", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "—", "Export. + Inf.", "Export. + Inf."],
    [1, "Comparación principal", "TensorRT INT8", "Desde cero", "Pesos aleatorios", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "—", "Export. + Inf.", "Export. + Inf."],
    [1, "Comparación principal", "TensorRT INT8", "Transfer learning", "COCO preentrenado", 640, 1, "v12, v26", "Segmentación", "N, S, M, L", "—", "Export. + Inf.", "Export. + Inf."],
    # Exp 2 — Tamaño de entrada
    [2, "Tamaño de entrada", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 320, 1, "v12, v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    [2, "Tamaño de entrada", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 1280, 1, "v12, v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    [2, "Tamaño de entrada", "PyTorch FP32", "Transfer learning", "COCO preentrenado", 320, 1, "v12, v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    [2, "Tamaño de entrada", "PyTorch FP32", "Transfer learning", "COCO preentrenado", 1280, 1, "v12, v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    # Exp 3 — Rendimiento por lote
    [3, "Rendimiento por lote", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 640, 4, "v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    [3, "Rendimiento por lote", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 640, 8, "v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    [3, "Rendimiento por lote", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 640, 16, "v26", "Segmentación", "N, S, M, L", "Inferencia", "Inferencia", "Inferencia"],
    # Exp 4 — Detección vs Segmentación
    [4, "Det. vs Seg.", "PyTorch FP32", "Desde cero", "Pesos aleatorios", 640, 1, "v12, v26", "Detección", "N, S, M, L", "Entren. + Inf.", "Inferencia", "Inferencia"],
    [4, "Det. vs Seg.", "PyTorch FP32", "Transfer learning", "COCO preentrenado", 640, 1, "v12, v26", "Detección", "N, S, M, L", "Entren. + Inf.", "Inferencia", "Inferencia"],
]

for r, row_data in enumerate(rows, 2):
    exp_num = row_data[0]
    fill = exp_fills[exp_num]
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = fill

# Anchos de columna
col_widths = [13, 22, 16, 18, 20, 9, 7, 14, 15, 14, 16, 16, 16]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

# ══════════════════════════════════════════════════════════════════════
# HOJA 2: Métricas
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Métricas")

metric_headers = ["Categoría", "Métrica", "Descripción"]
for col, h in enumerate(metric_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

metrics = [
    ["Velocidad", "Preprocesamiento", "Carga y normalización de imagen (ms/img)"],
    ["Velocidad", "Inferencia", "Pasada hacia adelante (ms/img)"],
    ["Velocidad", "Postprocesamiento", "NMS y generación de máscaras (ms/img)"],
    ["Velocidad", "Latencia total", "Suma de las tres anteriores (ms/img)"],
    ["Velocidad", "FPS", "Fotogramas por segundo"],
    ["Precisión", "mAP50", "Precisión media a IoU 0.50"],
    ["Precisión", "mAP50-95", "Precisión media promediada de IoU 0.50 a 0.95"],
    ["Eficiencia", "Vatios", "Consumo energético durante la inferencia (W)"],
    ["Eficiencia", "FPS/vatio", "Eficiencia energética (solo Jetsons, vía jtop)"],
]

for r, row_data in enumerate(metrics, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.alignment = left if c == 3 else center
        cell.border = thin_border

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 50

# ══════════════════════════════════════════════════════════════════════
# HOJA 3: Resumen de ejecuciones
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Resumen de Ejecuciones")

summary_headers = [
    "Experimento", "Enfoque",
    "Entrenamiento\n(RTX 5090)", "Inferencia\nRTX 5090",
    "Inferencia\nOrin AGX", "Inferencia\nOrin Nano",
    "Total\ninferencia",
]
for col, h in enumerate(summary_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Exp 1: 6 rows × 4 sizes × 2 archs = 48/Jetson; RTX only PyTorch rows: 2 × 4 × 2 = 16
# Exp 2: 4 rows × 4 sizes × 2 archs = 32/device
# Exp 3: 3 rows × 4 sizes × 1 arch = 12/device
# Exp 4: 1 row × 4 sizes × 2 archs = 8/device
summary_rows = [
    [1, "Comparación principal", 16, 16, 48, 48, 112],
    [2, "Tamaño de entrada", 0, 32, 32, 32, 96],
    [3, "Rendimiento por lote", 0, 12, 12, 12, 36],
    [4, "Det. vs Seg.", 16, 16, 16, 16, 48],
    ["", "TOTAL", 32, 76, 108, 108, 292],
]

for r, row_data in enumerate(summary_rows, 2):
    is_total = row_data[0] == ""
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = bold_font if is_total else normal_font
        cell.alignment = center
        cell.border = thin_border

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 16
ws3.column_dimensions["F"].width = 16
ws3.column_dimensions["G"].width = 14

# ── Guardar ─────────────────────────────────────────────────────────
output_path = "d:/tmp/TFM/ICM/BenchMarks/plan_benchmark.xlsx"
wb.save(output_path)
print(f"Archivo guardado en: {output_path}")
